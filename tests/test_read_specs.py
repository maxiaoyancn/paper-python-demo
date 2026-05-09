from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from scripts.generate_excel_random_data import (
    GroupSpec,
    RowSpec,
    _find_decimals_col,
    read_specs,
)

FIXTURES = Path(__file__).parent / "fixtures"


def test_find_decimals_col_k2():
    wb = openpyxl.load_workbook(FIXTURES / "sample-k2.xlsx", data_only=True)
    assert _find_decimals_col(wb.active) == 8


def test_find_decimals_col_k3():
    wb = openpyxl.load_workbook(FIXTURES / "sample-k3.xlsx", data_only=True)
    assert _find_decimals_col(wb.active) == 11


def test_find_decimals_col_missing(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["指标", "N1", "μ1", "σ1", "N2", "μ2", "σ2"])
    p = tmp_path / "no_decimals.xlsx"
    wb.save(p)
    wb2 = openpyxl.load_workbook(p, data_only=True)
    with pytest.raises(LookupError):
        _find_decimals_col(wb2.active)


def test_read_specs_k2():
    specs = read_specs(FIXTURES / "sample-k2.xlsx", sheet=None)
    assert len(specs) == 3  # 体重 / 身高 / 血糖
    spec = specs[0]
    assert isinstance(spec, RowSpec)
    assert spec.metric == "体重"
    assert spec.row_index == 2
    assert spec.decimals == 4
    assert len(spec.groups) == 2
    assert spec.groups[0] == GroupSpec(name="G1", n=10, mean=120.6263, sd=10.3698)
    assert spec.groups[1] == GroupSpec(name="G2", n=12, mean=110.36, sd=9.8635)


def test_read_specs_k3():
    specs = read_specs(FIXTURES / "sample-k3.xlsx", sheet=None)
    assert len(specs) == 2
    spec = specs[0]
    assert spec.metric == "体重"
    assert spec.decimals == 4
    assert len(spec.groups) == 3
    assert spec.groups[0].n == 8
    assert spec.groups[2].mean == 80.0


def test_read_specs_skips_invalid_rows(tmp_path, caplog):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["指标", "N1", "μ1", "σ1", "N2", "μ2", "σ2", "原始数据小数点后位数"])
    ws.append(["ok", 10, 100.0, 5.0, 10, 90.0, 5.0, 4])
    ws.append(["bad-N", 1, 100.0, 5.0, 10, 90.0, 5.0, 4])
    ws.append(["bad-SD", 10, 100.0, 0.0, 10, 90.0, 5.0, 4])
    ws.append(["bad-dec", 10, 100.0, 5.0, 10, 90.0, 5.0, -1])
    ws.append(["miss-mean", 10, None, 5.0, 10, 90.0, 5.0, 4])
    p = tmp_path / "mixed.xlsx"
    wb.save(p)

    with caplog.at_level("WARNING"):
        specs = read_specs(p, sheet=None)
    assert [s.metric for s in specs] == ["ok"]
    assert sum(1 for r in caplog.records if r.levelname == "WARNING") >= 4


def test_read_specs_no_decimals_col_raises(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["指标", "N1", "μ1", "σ1", "N2", "μ2", "σ2"])
    ws.append(["x", 10, 1.0, 1.0, 10, 1.0, 1.0])
    p = tmp_path / "no_dec.xlsx"
    wb.save(p)
    with pytest.raises(LookupError):
        read_specs(p, sheet=None)


def test_read_specs_k_gt_26_raises(tmp_path):
    """K=27 (decimals_col=83 = '原始数据小数点后位数') should be rejected."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["指标"]
    for i in range(27):
        headers.extend([f"G{i + 1} N", f"G{i + 1} mean", f"G{i + 1} SD"])
    headers.append("原始数据小数点后位数")
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    row = ["x"]
    for _ in range(27):
        row.extend([5, 1.0, 1.0])
    row.append(4)
    for c, v in enumerate(row, start=1):
        ws.cell(row=2, column=c).value = v
    p = tmp_path / "k27.xlsx"
    wb.save(p)
    with pytest.raises(LookupError, match="超过 26"):
        read_specs(p, sheet=None)

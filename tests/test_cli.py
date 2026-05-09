from __future__ import annotations

import shutil
from itertools import combinations
from pathlib import Path

import numpy as np
import openpyxl
import pytest
from openpyxl.utils import column_index_from_string
from scipy import stats as sp_stats

from scripts.generate_excel_random_data import LEVENE_ALPHA, main

ROOT = Path(__file__).resolve().parent.parent
FIXTURES = Path(__file__).parent / "fixtures"
SAMPLE_K2 = FIXTURES / "sample-k2.xlsx"
SAMPLE_K3 = FIXTURES / "sample-k3.xlsx"


def test_cli_requires_input_argument(capsys):
    with pytest.raises(SystemExit) as exc_info:
        main([])
    assert exc_info.value.code == 2
    assert "--input" in capsys.readouterr().err


def _clear_row1_after_inputs(path: Path, decimals_col: int) -> None:
    """Clear row 1 cells beyond the input area so the script writes fresh v3 headers.

    Without this, stale v1/v2 stat headers in the fixture would be preserved by the
    'non-empty header is kept' policy, masking v3 header behavior.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for c in range(decimals_col + 1, ws.max_column + 1):
        ws.cell(row=1, column=c).value = None
    wb.save(path)


def test_cli_writes_expected_layout_k2(tmp_path: Path):
    src = tmp_path / "in.xlsx"
    shutil.copy(SAMPLE_K2, src)
    _clear_row1_after_inputs(src, decimals_col=8)  # K=2: decimals_col=H=8
    rc = main(["--input", str(src), "--output", str(src), "--seed", "42"])
    assert rc == 0

    wb_in = openpyxl.load_workbook(SAMPLE_K2, data_only=True)
    wb_out = openpyxl.load_workbook(src, data_only=True)
    ws_in = wb_in.active
    ws_out = wb_out.active

    # A-H (1..8) preserved
    for r in range(1, ws_in.max_row + 1):
        for c in range(1, 9):
            assert (
                ws_out.cell(row=r, column=c).value == ws_in.cell(row=r, column=c).value
            ), (r, c)

    col_J = column_index_from_string("J")
    col_S = column_index_from_string("S")
    col_T = column_index_from_string("T")
    col_U = column_index_from_string("U")
    col_AF = column_index_from_string("AF")
    col_AL = column_index_from_string("AL")  # Levene p

    # New letter-based data headers
    assert ws_out.cell(row=1, column=col_J).value == "A1"
    assert ws_out.cell(row=1, column=col_S).value == "A10"
    assert ws_out.cell(row=1, column=col_U).value == "B1"
    assert ws_out.cell(row=1, column=col_AF).value == "B12"

    # Stat headers row 1
    assert ws_out.cell(row=1, column=col_AL).value == "Levene p"
    assert ws_out.cell(row=1, column=col_AL + 1).value == "是否方差齐"
    assert ws_out.cell(row=1, column=col_AL + 2).value == "Shapiro-Wilk min p"
    assert ws_out.cell(row=1, column=col_AL + 3).value == "是否正态"

    for r in range(2, 5):
        assert ws_out.cell(row=r, column=col_T).value is None  # blank between groups

    # Verify row 2 stats (体重)
    g1 = np.asarray(
        [ws_out.cell(row=2, column=c).value for c in range(col_J, col_J + 10)]
    )
    g2 = np.asarray(
        [ws_out.cell(row=2, column=c).value for c in range(col_U, col_U + 12)]
    )

    expected_levene = float(sp_stats.levene(g1, g2, center="median").pvalue)
    expected_sw = min(
        float(sp_stats.shapiro(g1).pvalue), float(sp_stats.shapiro(g2).pvalue)
    )
    equal_var = expected_levene > LEVENE_ALPHA
    if equal_var:
        expected_overall = float(sp_stats.f_oneway(g1, g2).pvalue)
    else:
        expected_overall = float(sp_stats.kruskal(g1, g2).pvalue)

    # col offsets: 0=Levene, 1=levene_flag, 2=SW, 3=normality_flag, 4=overall
    assert ws_out.cell(row=2, column=col_AL).value == pytest.approx(
        round(expected_levene, 4)
    )
    assert ws_out.cell(row=2, column=col_AL + 1).value in ("Y", "N")
    assert ws_out.cell(row=2, column=col_AL + 2).value == pytest.approx(
        round(expected_sw, 4)
    )
    assert ws_out.cell(row=2, column=col_AL + 3).value in ("Y", "N")
    assert ws_out.cell(row=2, column=col_AL + 4).value == pytest.approx(
        round(expected_overall, 4)
    )


def test_cli_writes_expected_layout_k3(tmp_path: Path):
    src = tmp_path / "in.xlsx"
    shutil.copy(SAMPLE_K3, src)
    _clear_row1_after_inputs(src, decimals_col=11)  # K=3: decimals_col=K=11
    rc = main(["--input", str(src), "--output", str(src), "--seed", "42"])
    assert rc == 0

    wb_in = openpyxl.load_workbook(SAMPLE_K3, data_only=True)
    wb = openpyxl.load_workbook(src, data_only=True)
    ws_in = wb_in.active
    ws = wb.active

    # A-K (1..11) preserved
    for r in range(1, ws_in.max_row + 1):
        for c in range(1, 12):
            assert ws.cell(row=r, column=c).value == ws_in.cell(row=r, column=c).value

    col_T = column_index_from_string("T")
    col_AE = column_index_from_string("AE")
    col_AR = column_index_from_string("AR")
    assert ws.cell(row=1, column=col_T).value == "A8"
    assert ws.cell(row=1, column=col_AE).value == "B10"
    assert ws.cell(row=1, column=col_AR).value == "C12"

    # K=3 stat block: stat_start=46, mean_sd=46-51,
    # levene=52, levene_flag=53, SW=54, normality_flag=55,
    # overall=56, raw=57-59, Q=60-62
    assert ws.cell(row=1, column=53).value == "是否方差齐"
    assert ws.cell(row=1, column=55).value == "是否正态"
    assert ws.cell(row=1, column=57).value == "A-B raw p"
    assert ws.cell(row=1, column=58).value == "A-C raw p"
    assert ws.cell(row=1, column=59).value == "B-C raw p"
    assert ws.cell(row=1, column=60).value == "A-B Q-value"

    raw_start = 57
    q_start = 60
    pair_count = 3

    g_starts_lens = [(13, 8), (22, 10), (33, 12)]
    groups = [
        np.asarray([ws.cell(row=2, column=c).value for c in range(start, start + n)])
        for (start, n) in g_starts_lens
    ]

    pairs = list(combinations(range(3), 2))
    levene_p = float(sp_stats.levene(*groups, center="median").pvalue)
    if levene_p > LEVENE_ALPHA:
        sw_min = min(float(sp_stats.shapiro(g).pvalue) for g in groups)
        if sw_min > LEVENE_ALPHA:
            matrix = sp_stats.tukey_hsd(*groups).pvalue
            expected_raw = [float(matrix[i][j]) for i, j in pairs]
            expected_q: list[float | None] = [None] * pair_count
        else:
            expected_raw = [
                float(sp_stats.ttest_ind(groups[i], groups[j], equal_var=False).pvalue)
                for i, j in pairs
            ]
            expected_q = [min(1.0, r * pair_count) for r in expected_raw]
    else:
        expected_raw = [
            float(sp_stats.ttest_ind(groups[i], groups[j], equal_var=False).pvalue)
            for i, j in pairs
        ]
        expected_q = [min(1.0, r * pair_count) for r in expected_raw]

    for i, exp in enumerate(expected_raw):
        got = ws.cell(row=2, column=raw_start + i).value
        assert got == pytest.approx(round(exp, 4)), (i, got, exp)
    for i, exp in enumerate(expected_q):
        got = ws.cell(row=2, column=q_start + i).value
        if exp is None:
            assert got is None
        else:
            assert got == pytest.approx(round(exp, 4))


def test_cli_seed_reproducible(tmp_path: Path):
    src1 = tmp_path / "a.xlsx"
    src2 = tmp_path / "b.xlsx"
    shutil.copy(SAMPLE_K2, src1)
    shutil.copy(SAMPLE_K2, src2)
    main(["--input", str(src1), "--output", str(src1), "--seed", "42"])
    main(["--input", str(src2), "--output", str(src2), "--seed", "42"])
    wb1 = openpyxl.load_workbook(src1, data_only=True).active
    wb2 = openpyxl.load_workbook(src2, data_only=True).active
    for r in range(1, wb1.max_row + 1):
        for c in range(1, wb1.max_column + 1):
            assert wb1.cell(row=r, column=c).value == wb2.cell(row=r, column=c).value, (
                r,
                c,
            )


def test_cli_no_decimals_col_returns_3(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["指标", "N1", "μ1", "σ1", "N2", "μ2", "σ2"])
    ws.append(["x", 10, 1.0, 1.0, 10, 1.0, 1.0])
    p = tmp_path / "no_dec.xlsx"
    wb.save(p)
    rc = main(["--input", str(p), "--output", str(p)])
    assert rc == 3


def test_cli_flags_consistent_with_p_values(tmp_path: Path):
    """For every data row, levene_flag/normality_flag must match their p-value cells."""
    src = tmp_path / "in.xlsx"
    shutil.copy(SAMPLE_K2, src)
    _clear_row1_after_inputs(src, decimals_col=8)
    rc = main(["--input", str(src), "--output", str(src), "--seed", "42"])
    assert rc == 0

    ws = openpyxl.load_workbook(src, data_only=True).active
    col_AL = column_index_from_string("AL")  # Levene p

    for r in range(2, 5):
        levene_p = ws.cell(row=r, column=col_AL).value
        levene_flag = ws.cell(row=r, column=col_AL + 1).value
        sw_p = ws.cell(row=r, column=col_AL + 2).value
        normality_flag = ws.cell(row=r, column=col_AL + 3).value

        expected_levene_flag = (
            "Y" if (levene_p is not None and levene_p > 0.05) else "N"
        )
        expected_normality_flag = "Y" if (sw_p is not None and sw_p > 0.05) else "N"

        assert levene_flag == expected_levene_flag, (r, levene_p, levene_flag)
        assert normality_flag == expected_normality_flag, (r, sw_p, normality_flag)

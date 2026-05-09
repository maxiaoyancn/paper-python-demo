"""Build K=3 / K=4 fixture xlsx files for v2 tests.

Layout reminder (K groups):
    A1=指标; B-D=组1 (N,μ,σ); E-G=组2; H-J=组3 (when K≥3); ... ;
    decimals_col = col(2 + 3K); decimals_col + 1 = remarks (optional);
    decimals_col + 2 onward = data area (script writes here).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def build_k3(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "指标",
        "G1 N",
        "G1 mean",
        "G1 SD",
        "G2 N",
        "G2 mean",
        "G2 SD",
        "G3 N",
        "G3 mean",
        "G3 SD",
        "原始数据小数点后位数",
        "备注（脚本不读）",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    rows = [
        ("体重", 8, 70.0, 5.0, 10, 75.0, 6.0, 12, 80.0, 7.0, 4),
        ("身高", 8, 165.0, 8.0, 10, 170.0, 9.0, 12, 175.0, 10.0, 2),
    ]
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v
    wb.save(path)


def build_k4(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "指标",
        "G1 N",
        "G1 mean",
        "G1 SD",
        "G2 N",
        "G2 mean",
        "G2 SD",
        "G3 N",
        "G3 mean",
        "G3 SD",
        "G4 N",
        "G4 mean",
        "G4 SD",
        "原始数据小数点后位数",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    rows = [("血糖", 6, 5.0, 1.0, 8, 6.0, 1.2, 10, 7.0, 1.5, 12, 8.0, 1.8, 3)]
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v
    wb.save(path)


if __name__ == "__main__":
    here = Path(__file__).parent
    build_k3(here / "sample-k3.xlsx")
    build_k4(here / "sample-k4.xlsx")
    print("ok")

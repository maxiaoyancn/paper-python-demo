"""按目标统计量为 Excel 中每行指标生成两组正态分布原始数据，并回算 mean/SD/p 值。"""

from __future__ import annotations

import argparse
import itertools
import logging
import secrets
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy import stats as sp_stats

logger = logging.getLogger(__name__)

LEVENE_ALPHA = 0.05
SD_TOLERANCE = 0.10  # ±10%
MAX_RETRY = 5
STAT_DIGITS = 4  # 统计回算列固定 4 位小数
LEVENE_HEADER = "Levene p"
LEVENE_FLAG_HEADER = "是否方差齐"
SHAPIRO_HEADER = "Shapiro-Wilk min p"
NORMALITY_FLAG_HEADER = "是否正态"
OVERALL_HEADER = "整体 p（ANOVA/KW）"


@dataclass(slots=True, frozen=True)
class GroupSpec:
    name: str
    n: int
    mean: float
    sd: float


@dataclass(slots=True, frozen=True)
class RowSpec:
    row_index: int
    metric: str
    groups: tuple[GroupSpec, ...]
    decimals: int


def _letter(i: int) -> str:
    """1 → 'A', 2 → 'B', ..., 26 → 'Z'. Raises ValueError for i outside 1..26."""
    if i < 1 or i > 26:
        raise ValueError(f"组序号必须在 1..26 内，收到 {i}")
    return chr(ord("A") + i - 1)


def _find_decimals_col(ws: Worksheet) -> int:
    """Scan row 1 headers; return the first column whose value contains '位数'."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and "位数" in v:
            return c
    raise LookupError("未在第 1 行找到包含'位数'的表头列")


def _coerce_int(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _coerce_float(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def read_specs(path: Path, sheet: str | None) -> list[RowSpec]:
    wb = load_workbook(path, data_only=True)
    ws: Worksheet = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    decimals_col = _find_decimals_col(ws)
    if (decimals_col - 2) % 3 != 0:
        raise LookupError(
            f"位数列在第 {decimals_col} 列，与 K 组三元组布局不兼容"
            "（需满足 (col-2) % 3 == 0）"
        )
    k = (decimals_col - 2) // 3
    if k < 2:
        raise LookupError(f"组数 K={k} < 2，至少需要两组数据")
    if k > 26:
        raise LookupError(f"组数 K={k} 超过 26 上限（字母 A..Z 范围）")

    out: list[RowSpec] = []
    for row_idx in range(2, ws.max_row + 1):
        metric = ws.cell(row=row_idx, column=1).value

        if all(
            ws.cell(row=row_idx, column=c).value is None
            for c in range(1, decimals_col + 1)
        ):
            continue

        problems: list[str] = []
        if not isinstance(metric, str) or not metric.strip():
            problems.append("缺少指标名")

        groups: list[GroupSpec] = []
        for i in range(k):
            base = 2 + 3 * i
            n = _coerce_int(ws.cell(row=row_idx, column=base).value)
            mean = _coerce_float(ws.cell(row=row_idx, column=base + 1).value)
            sd = _coerce_float(ws.cell(row=row_idx, column=base + 2).value)
            if n is None or n <= 1:
                problems.append(f"G{i + 1} 的 N 必须是 ≥2 的整数")
            if mean is None:
                problems.append(f"G{i + 1} 的均值缺失或非数值")
            if sd is None or sd <= 0:
                problems.append(f"G{i + 1} 的 SD 必须是 >0 的数值")
            if (
                n is not None
                and n > 1
                and mean is not None
                and sd is not None
                and sd > 0
            ):
                groups.append(GroupSpec(name=f"G{i + 1}", n=n, mean=mean, sd=sd))

        decimals = _coerce_int(ws.cell(row=row_idx, column=decimals_col).value)
        if decimals is None or decimals < 0:
            problems.append("decimals 缺失或为负")

        if problems:
            logger.warning(
                "跳过第 %d 行（%r）：%s", row_idx, metric, "；".join(problems)
            )
            continue

        out.append(
            RowSpec(
                row_index=row_idx,
                metric=metric,  # type: ignore[arg-type]
                groups=tuple(groups),
                decimals=decimals,  # type: ignore[arg-type]
            )
        )
    return out


def generate_one_group(
    target_mean: float,
    target_sd: float,
    size: int,
    rng: np.random.Generator,
    decimals: int,
) -> np.ndarray:
    """采样标准正态后做线性变换严格匹配 target，再按 decimals round。"""
    if size < 2:
        raise ValueError(f"size 必须 ≥2，收到 {size}")
    if target_sd <= 0:
        raise ValueError(f"target_sd 必须 >0，收到 {target_sd}")
    if decimals < 0:
        raise ValueError(f"decimals 必须 ≥0，收到 {decimals}")
    z = rng.standard_normal(size)
    z = (z - z.mean()) / z.std(ddof=1)
    x = target_mean + target_sd * z
    return np.round(x, decimals)


def generate_with_retry(
    metric: str,
    target_mean: float,
    target_sd: float,
    size: int,
    rng: np.random.Generator,
    decimals: int,
) -> np.ndarray:
    """在 SD ±10% 容差内重试，最多 MAX_RETRY 次，仍失败抛 RuntimeError。"""
    last_actual_sd: float | None = None
    for attempt in range(1, MAX_RETRY + 1):
        seed = int(rng.integers(0, 2**32 - 1))
        sub_rng = np.random.default_rng(seed)
        x = generate_one_group(target_mean, target_sd, size, sub_rng, decimals)
        actual_sd = float(np.std(x, ddof=1))
        last_actual_sd = actual_sd
        rel_err = abs(actual_sd - target_sd) / target_sd
        logger.debug(
            "metric=%s attempt=%d size=%d decimals=%d target_sd=%.6f actual_sd=%.6f rel_err=%.4f",
            metric,
            attempt,
            size,
            decimals,
            target_sd,
            actual_sd,
            rel_err,
        )
        if rel_err <= SD_TOLERANCE:
            return x
    raise RuntimeError(
        f"无法在容差内还原 SD: metric={metric}, target_sd={target_sd}, "
        f"last_actual_sd={last_actual_sd}, retries={MAX_RETRY}"
    )


def compute_levene(
    groups: list[np.ndarray], alpha: float = LEVENE_ALPHA
) -> tuple[float, bool]:
    """Brown-Forsythe Levene with center='median'. Returns (p, equal_var)."""
    res = sp_stats.levene(*groups, center="median")
    p = float(res.pvalue)
    if np.isnan(p):
        return float("nan"), False
    return p, p > alpha


def compute_shapiro_min(
    groups: list[np.ndarray], alpha: float = LEVENE_ALPHA
) -> tuple[float, bool]:
    """Per-group Shapiro-Wilk; return (min p, all_normal)."""
    ps: list[float] = []
    for i, g in enumerate(groups, start=1):
        if len(g) < 3:
            logger.warning(
                "G%d 样本量 N=%d < 3，Shapiro-Wilk 视为 0（不正态）", i, len(g)
            )
            ps.append(0.0)
            continue
        try:
            res = sp_stats.shapiro(g)
            p = float(res.pvalue)
            if np.isnan(p):
                p = 0.0
            ps.append(p)
        except Exception as e:  # noqa: BLE001
            logger.warning("G%d Shapiro-Wilk 失败: %s，视为 0", i, e)
            ps.append(0.0)
    min_p = min(ps)
    return min_p, min_p > alpha


def compute_overall(groups: list[np.ndarray], equal_var: bool) -> tuple[float, str]:
    """ANOVA when equal_var else Kruskal-Wallis. Returns (p, label)."""
    if equal_var:
        res = sp_stats.f_oneway(*groups)
        return float(res.pvalue), "ANOVA"
    res = sp_stats.kruskal(*groups)
    return float(res.pvalue), "KW"


def compute_pairwise(
    groups: list[np.ndarray], equal_var: bool, all_normal: bool
) -> tuple[list[float], list[float | None], str]:
    """Tukey HSD when equal_var+all_normal else Welch+Bonferroni.

    Returns (raw_ps, q_values, label). q_values are all None for the Tukey branch.
    """
    k = len(groups)
    pairs = list(itertools.combinations(range(k), 2))
    if equal_var and all_normal:
        matrix = sp_stats.tukey_hsd(*groups).pvalue
        raw = [float(matrix[i][j]) for i, j in pairs]
        q: list[float | None] = [None] * len(pairs)
        return raw, q, "Tukey"
    raw = [
        float(sp_stats.ttest_ind(groups[i], groups[j], equal_var=False).pvalue)
        for i, j in pairs
    ]
    n_pairs = len(pairs)
    q = [min(1.0, r * n_pairs) for r in raw]
    return raw, q, "Welch+Bonferroni"


@dataclass(slots=True, frozen=True)
class StatColIndices:
    mean_sd_pairs: tuple[tuple[int, int], ...]
    levene: int
    levene_flag: int
    shapiro_min: int
    normality_flag: int
    overall: int
    pairwise_raw: tuple[int, ...]
    pairwise_q: tuple[int, ...]


@dataclass(slots=True, frozen=True)
class Layout:
    decimals_col: int
    group_cols: tuple[range, ...]
    stat_cols: StatColIndices


def compute_layout(decimals_col: int, group_sizes: tuple[int, ...]) -> Layout:
    if len(group_sizes) < 2:
        raise ValueError(f"组数必须 ≥2，收到 {len(group_sizes)}")
    if any(n < 2 for n in group_sizes):
        raise ValueError(f"每组 N 必须 ≥2，收到 {group_sizes}")

    data_start = decimals_col + 2
    group_cols: list[range] = []
    cursor = data_start
    for n in group_sizes:
        group_cols.append(range(cursor, cursor + n))
        cursor += n + 1  # n cells + 1 blank between groups
    # cursor now points 1 past the trailing blank — that is exactly stats_start.
    stats_start = cursor

    pair_cols: list[tuple[int, int]] = []
    c = stats_start
    for _ in group_sizes:
        pair_cols.append((c, c + 1))
        c += 2

    levene = c
    c += 1
    levene_flag = c
    c += 1
    shapiro = c
    c += 1
    normality_flag = c
    c += 1
    overall = c
    c += 1

    k = len(group_sizes)
    n_pairs = k * (k - 1) // 2
    raw = tuple(range(c, c + n_pairs))
    c += n_pairs
    q = tuple(range(c, c + n_pairs))

    return Layout(
        decimals_col=decimals_col,
        group_cols=tuple(group_cols),
        stat_cols=StatColIndices(
            mean_sd_pairs=tuple(pair_cols),
            levene=levene,
            levene_flag=levene_flag,
            shapiro_min=shapiro,
            normality_flag=normality_flag,
            overall=overall,
            pairwise_raw=raw,
            pairwise_q=q,
        ),
    )


def _write_group_data_headers(
    ws: Worksheet,
    group_cols: tuple[range, ...],
    group_sizes: tuple[int, ...],
) -> None:
    """第 1 行写入每组的列表头：<letter><idx>（如 A1, A2, ..., A10）."""
    for i, (cols, _n) in enumerate(zip(group_cols, group_sizes, strict=True), start=1):
        letter = _letter(i)
        for idx, col in enumerate(cols, start=1):
            ws.cell(row=1, column=col).value = f"{letter}{idx}"


def _write_stat_headers(ws: Worksheet, stat_cols: StatColIndices, k: int) -> None:
    """第 1 行补写统计列表头（仅当原表头为空）."""
    headers: list[tuple[int, str]] = []
    for i, (mu_col, sd_col) in enumerate(stat_cols.mean_sd_pairs, start=1):
        letter = _letter(i)
        headers.append((mu_col, f"{letter} 组均值"))
        headers.append((sd_col, f"{letter} 组SD值"))
    headers.append((stat_cols.levene, LEVENE_HEADER))
    headers.append((stat_cols.levene_flag, LEVENE_FLAG_HEADER))
    headers.append((stat_cols.shapiro_min, SHAPIRO_HEADER))
    headers.append((stat_cols.normality_flag, NORMALITY_FLAG_HEADER))
    headers.append((stat_cols.overall, OVERALL_HEADER))
    pair_labels = [
        f"{_letter(i + 1)}-{_letter(j + 1)}"
        for i, j in itertools.combinations(range(k), 2)
    ]
    for col, label in zip(stat_cols.pairwise_raw, pair_labels, strict=True):
        headers.append((col, f"{label} raw p"))
    for col, label in zip(stat_cols.pairwise_q, pair_labels, strict=True):
        headers.append((col, f"{label} Q-value"))
    for col, text in headers:
        if ws.cell(row=1, column=col).value in (None, ""):
            ws.cell(row=1, column=col).value = text


def _round_stat(v: float | None) -> float | None:
    if v is None:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    return round(float(v), STAT_DIGITS)


def write_row(
    ws: Worksheet,
    row: RowSpec,
    generated: list[np.ndarray],
    layout: Layout,
    levene_p: float,
    equal_var: bool,
    sw_min_p: float,
    all_normal: bool,
    overall_p: float,
    raw_ps: list[float],
    q_values: list[float | None],
) -> None:
    group_sizes = tuple(g.n for g in row.groups)
    _write_group_data_headers(ws, layout.group_cols, group_sizes)
    _write_stat_headers(ws, layout.stat_cols, len(row.groups))

    for cols, data in zip(layout.group_cols, generated, strict=True):
        for value, col in zip(data.tolist(), cols, strict=True):
            ws.cell(row=row.row_index, column=col).value = value

    actual_means = [float(np.mean(d)) for d in generated]
    actual_sds = [float(np.std(d, ddof=1)) for d in generated]
    for (mu_col, sd_col), mu, sd in zip(
        layout.stat_cols.mean_sd_pairs, actual_means, actual_sds, strict=True
    ):
        ws.cell(row=row.row_index, column=mu_col).value = _round_stat(mu)
        ws.cell(row=row.row_index, column=sd_col).value = _round_stat(sd)

    ws.cell(row=row.row_index, column=layout.stat_cols.levene).value = _round_stat(
        levene_p
    )
    ws.cell(row=row.row_index, column=layout.stat_cols.levene_flag).value = (
        "Y" if equal_var else "N"
    )
    ws.cell(row=row.row_index, column=layout.stat_cols.shapiro_min).value = _round_stat(
        sw_min_p
    )
    ws.cell(row=row.row_index, column=layout.stat_cols.normality_flag).value = (
        "Y" if all_normal else "N"
    )
    ws.cell(row=row.row_index, column=layout.stat_cols.overall).value = _round_stat(
        overall_p
    )
    for col, p in zip(layout.stat_cols.pairwise_raw, raw_ps, strict=True):
        ws.cell(row=row.row_index, column=col).value = _round_stat(p)
    for col, q in zip(layout.stat_cols.pairwise_q, q_values, strict=True):
        ws.cell(row=row.row_index, column=col).value = _round_stat(q)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="按 Excel 表格中的目标统计量生成两组正态分布原始数据并回写。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="输入 Excel 文件路径（必填，请显式传入完整或绝对路径）",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 Excel 文件路径（默认与 --input 相同，即就地修改）",
    )
    parser.add_argument(
        "--sheet", default=None, help="指定 sheet 名（默认第一个 sheet）"
    )
    parser.add_argument(
        "--seed", type=int, default=None, help="随机种子（不传则随机生成并打印）"
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="启用 DEBUG 级别日志"
    )
    return parser


def _process_rows(
    ws: Worksheet,
    specs: Iterable[RowSpec],
    rng: np.random.Generator,
    decimals_col: int,
) -> int:
    """对每行 spec 执行 generate → 4 个 stats → write，返回成功处理的行数。"""
    count = 0
    for spec in specs:
        group_sizes = tuple(g.n for g in spec.groups)
        layout = compute_layout(decimals_col, group_sizes)

        generated: list[np.ndarray] = []
        for i, g in enumerate(spec.groups, start=1):
            seed = int(rng.integers(0, 2**32 - 1))
            arr = generate_with_retry(
                metric=f"{spec.metric}/G{i}",
                target_mean=g.mean,
                target_sd=g.sd,
                size=g.n,
                rng=np.random.default_rng(seed),
                decimals=spec.decimals,
            )
            generated.append(arr)

        levene_p, equal_var = compute_levene(generated)
        sw_min_p, all_normal = compute_shapiro_min(generated)
        overall_p, overall_label = compute_overall(generated, equal_var)
        raw_ps, q_values, branch_label = compute_pairwise(
            generated, equal_var, all_normal
        )

        write_row(
            ws,
            spec,
            generated,
            layout,
            levene_p,
            equal_var,
            sw_min_p,
            all_normal,
            overall_p,
            raw_ps,
            q_values,
        )
        count += 1
        logger.info(
            "处理完成 row=%d metric=%s K=%d levene_p=%.4f all_normal=%s overall=%s(%.4f) pairwise=%s",
            spec.row_index,
            spec.metric,
            len(spec.groups),
            levene_p,
            all_normal,
            overall_label,
            overall_p,
            branch_label,
        )
    return count


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    seed = args.seed if args.seed is not None else secrets.randbits(32)
    logger.info("使用随机种子 seed=%d", seed)
    rng = np.random.default_rng(seed)

    input_path: Path = args.input
    output_path: Path = args.output if args.output is not None else input_path
    if not input_path.exists():
        logger.error("输入文件不存在: %s", input_path)
        return 2

    try:
        specs = read_specs(input_path, args.sheet)
    except LookupError as e:
        logger.error("无法解析输入文件结构: %s", e)
        return 3

    if not specs:
        logger.warning("没有有效行可处理，退出")
        return 0

    wb = load_workbook(input_path)
    ws: Worksheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    decimals_col = _find_decimals_col(ws)
    logger.info("识别到位数列在第 %d 列, K=%d", decimals_col, (decimals_col - 2) // 3)
    n = _process_rows(ws, specs, rng, decimals_col)
    wb.save(output_path)
    logger.info("已写入 %d 行 → %s", n, output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())

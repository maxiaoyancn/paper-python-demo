"""按目标统计量为 Excel 中每行指标生成两组正态分布原始数据，并回算 mean/SD/p 值。"""

from __future__ import annotations

import argparse
import logging
import secrets
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy import stats as sp_stats

logger = logging.getLogger(__name__)

GROUP1_START_COL = 9  # I 列
LEVENE_ALPHA = 0.05
SD_TOLERANCE = 0.10  # ±10%
MAX_RETRY = 5
ROUND_DIGITS = 4
STAT_HEADERS = ("第一组均值", "第一组SD值", "第二组均值", "第二组SD值", "两组的p值")


@dataclass(slots=True, frozen=True)
class GroupSpec:
    name: str
    n: int
    mean: float
    sd: float


@dataclass(slots=True, frozen=True)
class RowSpec:
    row_index: int
    metric: str
    group1: GroupSpec
    group2: GroupSpec


def _coerce_positive_int(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _coerce_float(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def read_specs(path: Path, sheet: str | None) -> list[RowSpec]:
    wb = load_workbook(path, data_only=True)
    ws: Worksheet = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    out: list[RowSpec] = []
    for row_idx in range(2, ws.max_row + 1):
        metric = ws.cell(row=row_idx, column=1).value
        n = _coerce_positive_int(ws.cell(row=row_idx, column=2).value)
        mean1 = _coerce_float(ws.cell(row=row_idx, column=3).value)
        sd1 = _coerce_float(ws.cell(row=row_idx, column=4).value)
        m = _coerce_positive_int(ws.cell(row=row_idx, column=5).value)
        mean2 = _coerce_float(ws.cell(row=row_idx, column=6).value)
        sd2 = _coerce_float(ws.cell(row=row_idx, column=7).value)

        if metric is None and n is None and mean1 is None:
            continue

        problems: list[str] = []
        if not isinstance(metric, str) or not metric.strip():
            problems.append("缺少指标名")
        for label, val in (("N", n), ("M", m)):
            if val is None or val <= 1:
                problems.append(f"{label} 必须是 ≥2 的整数")
        for label, val in (("第一组均值", mean1), ("第二组均值", mean2)):
            if val is None:
                problems.append(f"{label} 缺失或非数值")
        for label, val in (("第一组 SD", sd1), ("第二组 SD", sd2)):
            if val is None or val <= 0:
                problems.append(f"{label} 必须是 >0 的数值")
        if problems:
            logger.warning(
                "跳过第 %d 行（%r）：%s", row_idx, metric, "；".join(problems)
            )
            continue

        out.append(
            RowSpec(
                row_index=row_idx,
                metric=metric,
                group1=GroupSpec(name="第一组", n=n, mean=mean1, sd=sd1),
                group2=GroupSpec(name="第二组", n=m, mean=mean2, sd=sd2),
            )
        )
    return out


def generate_one_group(
    target_mean: float,
    target_sd: float,
    size: int,
    rng: np.random.Generator,
) -> np.ndarray:
    """采样标准正态后做线性变换，使未 round 前 mean/SD 严格匹配目标，再 round 到 4 位小数。"""
    if size < 2:
        raise ValueError(f"size 必须 ≥2，收到 {size}")
    if target_sd <= 0:
        raise ValueError(f"target_sd 必须 >0，收到 {target_sd}")
    z = rng.standard_normal(size)
    z = (z - z.mean()) / z.std(ddof=1)
    x = target_mean + target_sd * z
    return np.round(x, ROUND_DIGITS)


def generate_with_retry(
    metric: str,
    target_mean: float,
    target_sd: float,
    size: int,
    rng: np.random.Generator,
) -> np.ndarray:
    """在 SD ±10% 容差内重试，最多 MAX_RETRY 次，仍失败抛 RuntimeError。"""
    last_actual_sd: float | None = None
    for attempt in range(1, MAX_RETRY + 1):
        seed = int(rng.integers(0, 2**32 - 1))
        sub_rng = np.random.default_rng(seed)
        x = generate_one_group(target_mean, target_sd, size, sub_rng)
        actual_sd = float(np.std(x, ddof=1))
        last_actual_sd = actual_sd
        rel_err = abs(actual_sd - target_sd) / target_sd
        logger.debug(
            "metric=%s attempt=%d size=%d target_sd=%.6f actual_sd=%.6f rel_err=%.4f",
            metric,
            attempt,
            size,
            target_sd,
            actual_sd,
            rel_err,
        )
        if rel_err <= SD_TOLERANCE:
            return x
    raise RuntimeError(
        f"无法在容差内还原 SD: metric={metric}, target_sd={target_sd}, "
        f"last_actual_sd={last_actual_sd}, retries={MAX_RETRY}"
    )


def compute_stats(
    g1: np.ndarray, g2: np.ndarray, alpha: float = LEVENE_ALPHA
) -> tuple[float, float, float, float, float, bool]:
    """返回 (mean1, sd1, mean2, sd2, p_value, equal_var)。"""
    mean1 = float(np.mean(g1))
    sd1 = float(np.std(g1, ddof=1))
    mean2 = float(np.mean(g2))
    sd2 = float(np.std(g2, ddof=1))
    levene_p = float(sp_stats.levene(g1, g2, center="median").pvalue)
    equal_var = levene_p >= alpha
    p_value = float(sp_stats.ttest_ind(g1, g2, equal_var=equal_var).pvalue)
    logger.debug(
        "compute_stats levene_p=%.4f equal_var=%s p_value=%.4f",
        levene_p,
        equal_var,
        p_value,
    )
    return mean1, sd1, mean2, sd2, p_value, equal_var


@dataclass(slots=True, frozen=True)
class Layout:
    group1_cols: range
    group2_cols: range
    stat_cols: range


def compute_layout(n: int, m: int) -> Layout:
    if n < 2 or m < 2:
        raise ValueError(f"N、M 必须 ≥2，收到 N={n}、M={m}")
    g1_start = GROUP1_START_COL
    g1_end = g1_start + n - 1
    g2_start = g1_end + 2
    g2_end = g2_start + m - 1
    stat_start = g2_end + 2
    stat_end = stat_start + 4
    return Layout(
        group1_cols=range(g1_start, g1_end + 1),
        group2_cols=range(g2_start, g2_end + 1),
        stat_cols=range(stat_start, stat_end + 1),
    )


def write_row(
    ws: Worksheet,
    row: RowSpec,
    g1: np.ndarray,
    g2: np.ndarray,
    stats: tuple[float, float, float, float, float, bool],
    layout: Layout,
) -> None:
    mean1, sd1, mean2, sd2, p_value, _equal_var = stats

    last_g1_col = layout.group1_cols.stop - 1
    for idx, col in enumerate(layout.group1_cols, start=1):
        if col == last_g1_col:
            ws.cell(row=1, column=col).value = f"{row.group1.n}（N）"
        else:
            ws.cell(row=1, column=col).value = idx

    last_g2_col = layout.group2_cols.stop - 1
    for idx, col in enumerate(layout.group2_cols, start=1):
        if col == last_g2_col:
            ws.cell(row=1, column=col).value = f"{row.group2.n}（M）"
        else:
            ws.cell(row=1, column=col).value = idx

    for header, col in zip(STAT_HEADERS, layout.stat_cols, strict=True):
        existing = ws.cell(row=1, column=col).value
        if existing in (None, ""):
            ws.cell(row=1, column=col).value = header

    for value, col in zip(g1.tolist(), layout.group1_cols, strict=True):
        ws.cell(row=row.row_index, column=col).value = value
    for value, col in zip(g2.tolist(), layout.group2_cols, strict=True):
        ws.cell(row=row.row_index, column=col).value = value

    stat_values = (mean1, sd1, mean2, sd2, p_value)
    for value, col in zip(stat_values, layout.stat_cols, strict=True):
        ws.cell(row=row.row_index, column=col).value = float(value)


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="按 Excel 表格中的目标统计量生成两组正态分布原始数据并回写。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="输入 Excel 文件路径（必填，请显式传入完整或绝对路径）",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 Excel 文件路径（默认与 --input 相同，即就地修改）",
    )
    parser.add_argument(
        "--sheet", default=None, help="指定 sheet 名（默认第一个 sheet）"
    )
    parser.add_argument(
        "--seed", type=int, default=None, help="随机种子（不传则随机生成并打印）"
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="启用 DEBUG 级别日志"
    )
    return parser


def _process_rows(
    ws: Worksheet, specs: Iterable[RowSpec], rng: np.random.Generator
) -> int:
    """对每行 spec 执行 generate→compute→write，返回成功处理的行数。"""
    count = 0
    for spec in specs:
        layout = compute_layout(spec.group1.n, spec.group2.n)
        seed1 = int(rng.integers(0, 2**32 - 1))
        seed2 = int(rng.integers(0, 2**32 - 1))
        g1 = generate_with_retry(
            metric=f"{spec.metric}/G1",
            target_mean=spec.group1.mean,
            target_sd=spec.group1.sd,
            size=spec.group1.n,
            rng=np.random.default_rng(seed1),
        )
        g2 = generate_with_retry(
            metric=f"{spec.metric}/G2",
            target_mean=spec.group2.mean,
            target_sd=spec.group2.sd,
            size=spec.group2.n,
            rng=np.random.default_rng(seed2),
        )
        stats_tuple = compute_stats(g1, g2)
        write_row(ws, spec, g1, g2, stats_tuple, layout)
        count += 1
        logger.info(
            "处理完成 row=%d metric=%s p_value=%.4f equal_var=%s",
            spec.row_index,
            spec.metric,
            stats_tuple[4],
            stats_tuple[5],
        )
    return count


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )

    seed = args.seed if args.seed is not None else secrets.randbits(32)
    logger.info("使用随机种子 seed=%d", seed)
    rng = np.random.default_rng(seed)

    input_path: Path = args.input
    output_path: Path = args.output if args.output is not None else input_path
    if not input_path.exists():
        logger.error("输入文件不存在: %s", input_path)
        return 2

    specs = read_specs(input_path, args.sheet)
    if not specs:
        logger.warning("没有有效行可处理，退出")
        return 0

    wb = load_workbook(input_path)
    ws: Worksheet = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    n = _process_rows(ws, specs, rng)
    wb.save(output_path)
    logger.info("已写入 %d 行 → %s", n, output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
